import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Border, Side, Font
import json
import sys

def analyze_sheet(sheet: Worksheet):
    """Analyze a worksheet and return its properties"""
    info = {
        "name": sheet.title,
        "dimensions": sheet.dimensions,
        "frozen_rows": None,
        "frozen_cols": None,
        "columns": [],
        "merged_cells": [],
        "data_validation": [],
        "conditional_formatting": bool(sheet.conditional_formatting),
        "row_count": sheet.max_row,
        "col_count": sheet.max_column
    }
    
    # Check frozen panes
    if sheet.freeze_panes:
        col, row = openpyxl.utils.cell.coordinate_from_string(sheet.freeze_panes)
        col_idx = openpyxl.utils.column_index_from_string(col)
        info["frozen_cols"] = col_idx - 1
        info["frozen_rows"] = row - 1
    
    # Get column information
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header_cell = sheet[f"{col_letter}1"]
        col_info = {
            "index": col_idx,
            "letter": col_letter,
            "header": header_cell.value if header_cell.value else f"Column {col_letter}"
        }
        info["columns"].append(col_info)
    
    # Get merged cells
    for merged_range in sheet.merged_cells.ranges:
        info["merged_cells"].append(str(merged_range))
    
    # Get data validation
    for range_addr, dv in sheet.data_validations.dataValidation:
        info["data_validation"].append({
            "range": range_addr,
            "type": dv.type,
            "formula1": dv.formula1 if hasattr(dv, "formula1") else None,
            "formula2": dv.formula2 if hasattr(dv, "formula2") else None
        })
    
    return info

def analyze_excel(filepath):
    """Analyze the Excel file and return information about its structure"""
    workbook = openpyxl.load_workbook(filepath)
    
    result = {
        "filename": filepath,
        "sheet_names": workbook.sheetnames,
        "sheets": {}
    }
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        result["sheets"][sheet_name] = analyze_sheet(sheet)
        
        # Get a sample of the data (first 10 rows)
        data_sample = []
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_value = cell.value
                
                # If cell has formatting, add that info
                cell_format = {}
                if cell.fill.patternType != 'none':
                    cell_format["fill"] = cell.fill.patternType
                    if hasattr(cell.fill, "fgColor") and cell.fill.fgColor.type:
                        cell_format["fill_color"] = cell.fill.fgColor.rgb or "default"
                
                if cell.font and (cell.font.bold or cell.font.italic or cell.font.size):
                    cell_format["font"] = {
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "size": cell.font.size,
                        "color": cell.font.color.rgb if cell.font.color else None
                    }
                
                row_data.append({
                    "value": str(cell_value) if cell_value is not None else None,
                    "format": cell_format if cell_format else None
                })
            data_sample.append(row_data)
        
        result["sheets"][sheet_name]["data_sample"] = data_sample
    
    return result

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python analyze_excel.py <excel_file>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    try:
        result = analyze_excel(filepath)
        print(json.dumps(result, indent=2))
    except Exception as e:
        print(f"Error analyzing Excel file: {e}", file=sys.stderr)
        sys.exit(1)