import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import json
import sys

def analyze_sheet(sheet):
    """Analyze a worksheet and return its properties"""
    info = {
        "name": sheet.title,
        "dimensions": sheet.dimensions,
        "frozen_rows": None,
        "frozen_cols": None,
        "columns": [],
        "merged_cells": [],
        "data_validation": [],
        "conditional_formatting": bool(sheet.conditional_formatting),
        "row_count": sheet.max_row,
        "col_count": sheet.max_column
    }
    
    # Check frozen panes
    if sheet.freeze_panes:
        cell = sheet.freeze_panes
        if cell:
            col, row = coordinate_from_string(cell)
            col_idx = column_index_from_string(col)
            info["frozen_cols"] = col_idx - 1
            info["frozen_rows"] = row - 1
    
    # Get column information
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header_cell = sheet[f"{col_letter}1"]
        col_info = {
            "index": col_idx,
            "letter": col_letter,
            "header": str(header_cell.value) if header_cell.value else f"Column {col_letter}"
        }
        info["columns"].append(col_info)
    
    # Get merged cells
    for merged_range in sheet.merged_cells.ranges:
        info["merged_cells"].append(str(merged_range))
    
    # Get data validation
    try:
        for coord in sheet.data_validations.dataValidation:
            for sqref in coord.sqref:
                info["data_validation"].append({
                    "range": str(sqref),
                    "type": coord.type
                })
    except (AttributeError, TypeError):
        pass  # Handle empty or different data validation structure
    
    # Get first few rows of data
    data_sample = []
    for row_idx in range(1, min(20, sheet.max_row + 1)):
        row_data = {}
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            cell = sheet.cell(row=row_idx, column=col_idx)
            row_data[col_letter] = str(cell.value) if cell.value is not None else None
        data_sample.append(row_data)
    
    info["data_sample"] = data_sample
    return info

def analyze_excel(filepath):
    """Analyze the Excel file and return information about its structure"""
    workbook = openpyxl.load_workbook(filepath)
    
    result = {
        "filename": filepath,
        "sheet_names": workbook.sheetnames,
        "sheets": {}
    }
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        result["sheets"][sheet_name] = analyze_sheet(sheet)
    
    return result

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python analyze_excel_simple.py <excel_file>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    try:
        result = analyze_excel(filepath)
        print(json.dumps(result, indent=2))
    except Exception as e:
        print(f"Error analyzing Excel file: {e}", file=sys.stderr)
        sys.exit(1)